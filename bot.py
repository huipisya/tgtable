import logging
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, MessageHandler, CallbackQueryHandler, ContextTypes, filters
import openpyxl
from openpyxl import Workbook
import os
import re # Для работы с регулярными выражениями
from datetime import datetime

# Настройка логирования
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# Имя директории для хранения файлов пользователей
DATA_DIR = os.getenv('DATA_DIR', 'user_data')
os.makedirs(DATA_DIR, exist_ok=True)

# ID чата для бэкапа
BACKUP_CHAT_ID = os.getenv('BACKUP_CHAT_ID')

# Глобальная переменная для application
app = None

# --- ФУНКЦИИ РАБОТЫ С ИНДИВИДУАЛЬНЫМИ ФАЙЛАМИ ---
# (Все функции get_user_excel_file, init_user_excel, send_backup_for_user остаются без изменений)

def get_user_excel_file(user_id: int) -> str:
    """Возвращает путь к Excel-файлу конкретного пользователя."""
    return os.path.join(DATA_DIR, f'user_{user_id}.xlsx')

def init_user_excel(user_id: int):
    """Инициализирует Excel-файл для конкретного пользователя, если его нет."""
    excel_file = get_user_excel_file(user_id)
    
    if not os.path.exists(excel_file):
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Посты"
        
        # Заголовки
        headers = ['№', 'Ссылка', 'Статус', 'Дата добавления']
        ws.append(headers)
        
        # Стиль для заголовков
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        alignment_center = Alignment(horizontal="center", vertical="center")
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Применяем стиль к заголовкам
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = alignment_center
            cell.border = thin_border
        
        # Устанавливаем ширину столбцов
        ws.column_dimensions['A'].width = 8   # №
        ws.column_dimensions['B'].width = 60  # Ссылка
        ws.column_dimensions['C'].width = 20  # Статус
        ws.column_dimensions['D'].width = 22  # Дата
        
        wb.save(excel_file)
        logger.info(f"Создан новый Excel файл для пользователя {user_id}")

def send_backup_for_user(user_id: int):
    """Отправить бэкап индивидуального файла пользователя в указанный чат."""
    if BACKUP_CHAT_ID:
        excel_file = get_user_excel_file(user_id)
        if os.path.exists(excel_file):
            try:
                import asyncio
                async def _send():
                    try:
                        with open(excel_file, 'rb') as f:
                            await app.bot.send_document(
                                chat_id=BACKUP_CHAT_ID,
                                document=f,
                                filename=f'backup_user_{user_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
                            )
                        logger.info(f"Бэкап для пользователя {user_id} отправлен успешно")
                    except Exception as e:
                        logger.error(f"Ошибка отправки бэкапа для пользователя {user_id}: {e}")

                asyncio.create_task(_send())
            except Exception as e:
                logger.error(f"Ошибка подготовки бэкапа для пользователя {user_id}: {e}")
        else:
            logger.warning(f"Файл для бэкапа пользователя {user_id} не найден.")


# --- ОБНОВЛЁННЫЕ ФУНКЦИИ РАБОТЫ С EXCEL ---
# (Функции get_next_number, add_post_to_excel, delete_post_from_excel, update_post_status остаются без изменений)

def get_next_number(user_id: int):
    excel_file = get_user_excel_file(user_id)
    if not os.path.exists(excel_file):
        init_user_excel(user_id)
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    return ws.max_row

def add_post_to_excel(user_id: int, link: str, status=None):
    from openpyxl.styles import Alignment, Border, Side
    
    excel_file = get_user_excel_file(user_id)
    if not os.path.exists(excel_file):
        init_user_excel(user_id)

    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    row = ws.max_row + 1
    number = row - 1
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    ws[f'A{row}'] = number
    ws[f'A{row}'].alignment = Alignment(horizontal="center", vertical="center")
    ws[f'A{row}'].border = thin_border
    
    ws[f'B{row}'].hyperlink = link
    ws[f'B{row}'].value = link
    ws[f'B{row}'].border = thin_border
    
    ws[f'C{row}'] = status if status else ""
    ws[f'C{row}'].alignment = Alignment(horizontal="center", vertical="center")
    ws[f'C{row}'].border = thin_border
    
    ws[f'D{row}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws[f'D{row}'].alignment = Alignment(horizontal="center", vertical="center")
    ws[f'D{row}'].border = thin_border
    
    wb.save(excel_file)
    send_backup_for_user(user_id)
    return number

def delete_post_from_excel(user_id: int, link: str):
    excel_file = get_user_excel_file(user_id)
    if not os.path.exists(excel_file):
        logger.warning(f"Попытка удаления из несуществующего файла для пользователя {user_id}")
        return False

    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    deleted = False
    
    for row in range(2, ws.max_row + 1):
        if ws[f'B{row}'].value == link:
            ws.delete_rows(row)
            for i in range(row, ws.max_row + 1):
                ws[f'A{i}'] = i - 1
            wb.save(excel_file)
            deleted = True
            break
            
    if deleted:
        send_backup_for_user(user_id)
    
    return deleted

def update_post_status(user_id: int, link: str, status: str):
    excel_file = get_user_excel_file(user_id)
    if not os.path.exists(excel_file):
        logger.warning(f"Попытка обновления статуса в несуществующем файле для пользователя {user_id}")
        return False

    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    
    for row in range(2, ws.max_row + 1):
        if ws[f'B{row}'].value == link:
            ws[f'C{row}'] = status
            wb.save(excel_file)
            return True
    return False

def get_export_button():
    keyboard = [[InlineKeyboardButton("📊 Отправить мою базу данных", callback_data='export_db')]]
    return InlineKeyboardMarkup(keyboard)

# --- НОВАЯ ФУНКЦИЯ ДЛЯ ИЗВЛЕЧЕНИЯ ССЫЛОК ---

def extract_telegram_link(text: str) -> str:
    """
    Извлекает первую попавшуюся ссылку на пост/канал в Telegram из текста.
    Поддерживает форматы: https://t.me/username/post_number, https://t.me/username
    """
    # Регулярное выражение для поиска Telegram-ссылок
    # https://t.me/username/12345 или https://t.me/username
    # (?i) - нечувствительность к регистру (опционально)
    # (?:...) - не захватывающая группа
    pattern = r'https?://(?:t\.me|telegram\.me)/(?:[a-zA-Z0-9_]+)(?:/[0-9]+)?(?:/[a-zA-Z0-9_]+)?'
    match = re.search(pattern, text)
    if match:
        return match.group(0)
    return ""

# --- ОБНОВЛЁННЫЕ ОБРАБОТЧИКИ ---

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    init_user_excel(user_id)
    await update.message.reply_text(
        f"👋 Привет, {update.effective_user.first_name}! Я бот для сохранения постов.\n\n"
        "Ты можешь:\n"
        "1. Просто *переслать* мне пост из Telegram.\n"
        "2. Отправить мне *ссылку* на пост в Telegram.\n\n"
        "Я сохраню её в *твою* персональную базу данных.\n\n"
        "Команды:\n"
        "/export - выгрузить *твою* базу данных в Excel\n"
        "/stats - статистика *твоих* постов"
    )

# Обработка текста и пересланных сообщений
async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    init_user_excel(user_id)

    # --- Проверяем пересланное сообщение ---
    forwarded_message = update.message.forward_from_message_id
    if forwarded_message:
        # Пытаемся получить информацию о пересланном сообщении, чтобы найти ссылку
        # К сожалению, Bot API не всегда предоставляет прямую ссылку на пересланное сообщение.
        # Но если сообщение переслано из канала или супергруппы с username, часто можно сформировать ссылку.
        # Более надёжный способ - это проверить, есть ли в *этом* сообщении (update.message) текст или caption,
        # и попытаться извлечь оттуда ссылку.
        # Однако, если *только* переслано, и нет текста в текущем сообщении, сложно.
        # Попробуем проверить текст/подпись текущего сообщения на наличие ссылки.
        # Это сработает, если пользователь переслал сообщение и добавил комментарий с ссылкой.
        # Для *чистой* пересылки без комментария, Telegram API не даёт прямой URL.
        # Поэтому, если в текущем сообщении нет текста/подписи, мы не сможем обработать *только* пересылку.
        # Чтобы обрабатывать *только* пересланные посты, нужно полагаться на то, что в них есть ссылка,
        # или использовать более сложные методы (например, получение истории канала по ID, что сложно и требует прав).
        # Учитывая стандартное поведение Telegram, часто ссылка появляется в теле пересланного сообщения или в подсказке.
        # Проверим, есть ли текст или подпись в *этом* (текущем) сообщении, и если да, извлечём оттуда ссылку.
        # Telegram часто вставляет автоматическую ссылку при пересылке, если она была в оригинале.
        # Но если пользователь пересылает *без* комментария, и в оригинале был только текст, а не встроенный URL,
        # то в текущем сообщении бота может не быть ссылки.
        # Вывод: обработка *чистой* пересылки без текста/ссылки в текущем сообщении через Bot API - ограничена.
        # Мы можем обрабатывать пересылку, если в *текущем* сообщении есть текст, содержащий ссылку.
        # Это покрывает большинство случаев, особенно если пользователь пересылает с комментарием или если Telegram автоматически добавляет ссылку.

        # Сначала проверим текст текущего сообщения
        text = update.message.text or update.message.caption or ""
        if text:
            link = extract_telegram_link(text)
            if link:
                # Сохраняем ссылку в контексте пользователя
                context.user_data['current_link'] = link
                # Вызываем логику обработки ссылки (ту же, что и для текста)
                await _present_link_options(update, context, link)
                return # Обработали, выходим

    # --- Проверяем текст или подпись текущего сообщения (не пересланного) ---
    text = update.message.text or update.message.caption or ""
    if text:
        link = extract_telegram_link(text)
        if link:
            # Сохраняем ссылку в контексте пользователя
            context.user_data['current_link'] = link
            await _present_link_options(update, context, link)
            return # Обработали, выходим

    # Если не переслано и нет текста/ссылки, сообщаем пользователю
    if not forwarded_message and not text:
        await update.message.reply_text("❌ Я не нашёл ссылку в твоём сообщении. Отправь ссылку или перешли пост с комментарием содержащим ссылку.")
    elif not forwarded_message and text and not extract_telegram_link(text):
        await update.message.reply_text("❌ Я не нашёл действительную ссылку на пост в Telegram в твоём сообщении.")


# Вспомогательная функция для представления опций после извлечения ссылки
async def _present_link_options(update: Update, context: ContextTypes.DEFAULT_TYPE, link: str):
    """Предлагает пользователю действия после извлечения ссылки."""
    keyboard = [
        [InlineKeyboardButton("📝 Добавить пост + статус", callback_data='add_with_status')],
        [InlineKeyboardButton("🗑️ Удалить пост", callback_data='delete_post')],
        [InlineKeyboardButton("✅ Просто добавить пост", callback_data='add_simple')]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await update.message.reply_text(
        f"📌 Пост получен!\n\nСсылка: {link}\n\nВыбери действие:",
        reply_markup=reply_markup
    )


# Обработка нажатий на кнопки (остаётся без изменений)
async def button_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    user_id = query.from_user.id

    if query.data == 'export_db':
        excel_file = get_user_excel_file(user_id)
        if os.path.exists(excel_file):
            await query.message.reply_document(
                document=open(excel_file, 'rb'),
                filename=f'my_posts_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            )
        else:
            await query.edit_message_text("❌ Твоя база данных пуста. Добавь хотя бы один пост.")
        return
    
    link = context.user_data.get('current_link')
    
    if not link:
        await query.edit_message_text("❌ Ошибка: ссылка не найдена. Отправь ссылку заново.")
        return
    
    if query.data == 'add_simple':
        number = add_post_to_excel(user_id, link)
        await query.edit_message_text(
            f"✅ Пост #{number} добавлен в *твою* базу данных!\n\n"
            f"Ссылка: {link}",
            reply_markup=get_export_button()
        )
        context.user_data.clear()
        
    elif query.data == 'add_with_status':
        context.user_data['waiting_for_status'] = True
        await query.edit_message_text(
            f"📝 Введи статус для поста:\n\n{link}\n\n"
            "Например: Одобрено, На проверке, Отклонено и т.д."
        )
        
    elif query.data == 'delete_post':
        success = delete_post_from_excel(user_id, link)
        if success:
            await query.edit_message_text(
                f"🗑️ Пост удалён из *твоей* базы данных!\n\n"
                f"Ссылка: {link}",
                reply_markup=get_export_button()
            )
        else:
            await query.edit_message_text(
                f"❌ Не удалось удалить пост. Возможно, его уже нет в базе.\n\n"
                f"Ссылка: {link}",
                reply_markup=get_export_button()
            )
        context.user_data.clear()

# Обработка текстовых сообщений (для статуса) и теперь также пересланных
# Убираем старый обработчик MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text)
# Заменяем его на более общий, который может обрабатывать и текст, и пересылки
# filters.TEXT означает сообщение с текстом
# filters.FORWARDED означает пересланное сообщение (любое)
# Мы хотим обрабатывать *любое* сообщение, в котором есть потенциальная ссылка
# filters.TEXT или filters.CAPTION (для медиа)
# filters.FORWARDED сам по себе не означает, что в текущем сообщении есть текст/ссылка
# Нам нужно проверять, есть ли *в текущем сообщении* текст или подпись (caption)
# Поэтому используем обработчик на всё, что содержит текст или подпись
# MessageHandler(filters.TEXT | filters.CAPTION & ~filters.COMMAND, handle_message)
# Однако filters.CAPTION сам по себе не является фильтром для MessageHandler.
# Нужно использовать filters.UpdateType.MESSAGE и проверять наличие caption или text внутри.
# Или использовать комбинацию фильтров, которые срабатывают, если *есть* текст или подпись.
# filters.TEXT | filters.PHOTO & filters.caption (например)
# Лучше всего будет использовать один обработчик на всё, что не команда, и проверять внутри.
# filters.TEXT срабатывает, если есть .text
# filters.PHOTO срабатывает, если есть .photo, и можно проверить .caption
# filters.VIDEO и т.д.
# filters.UpdateType.MESSAGE срабатывает всегда, когда приходит сообщение.
# Но фильтры типа filters.TEXT идут в приоритете.
# filters.TEXT означает наличие .text
# filters.CAPTION не существует как отдельный фильтр MessageHandler
# filters.PHOTO & filters.CAPTION_TEXT (новая версия python-telegram-bot) или нужно проверять вручную.
# В версии 20.8, можно использовать:
# filters.PHOTO & filters.Caption.TEXT
# filters.VIDEO & filters.Caption.TEXT
# Но проще и универсальнее проверить вручную внутри обработчика.
# Я буду использовать фильтр, который сработает, если есть .text ИЛИ .caption
# filters.TEXT или filters.PHOTO (и проверка caption внутри), или общий фильтр и проверка внутри.
# filters.TEXT | (filters.PHOTO & lambda u: hasattr(u.message, 'caption') and u.message.caption)
# filters.TEXT | filters.PHOTO | filters.VIDEO | ... и проверить внутри.
# Или просто:
# MessageHandler((filters.TEXT | filters.PHOTO | filters.VIDEO | filters.AUDIO | filters.DOCUMENT) & ~filters.COMMAND, handle_message)
# И внутри проверить text и caption.
# filters.TEXT уже включает в себя наличие .text
# filters.PHOTO включает наличие .photo
# filters.CAPTION_TEXT (новый синтаксис) означает, что у сообщения есть подпись (caption), и она содержит текст.
# filters.TEXT
# filters.CAPTION = lambda u: u.effective_message.caption is not None
# filters.CAPTION_TEXT = filters.CAPTION & filters.TEXT
# filters.TEXT означает, что в сообщении есть поле text (и оно не None)
# filters.CAPTION означает, что в сообщении есть поле caption (и оно не None)
# filters.CAPTION_TEXT означает, что в сообщении есть поле caption И оно содержит текст (не пустое)
# filters.TEXT | filters.CAPTION_TEXT (или что-то подобное)
# filters.TEXT | (filters.CAPTION & filters.TEXT) # caption проверяет наличие caption, TEXT внутри проверяет его содержимое
# filters.TEXT | filters.Caption.TEXT # (если Caption регистрозависим)
# filters.TEXT | filters.CAPTION.TEXT # (CAPTION)
# filters.TEXT | filters.Caption.TEXT # (Caption)
# Проверим документацию для конкретной версии.
# В версии 20.8:
# filters.TEXT
# filters.CAPTION
# filters.CAPTION.TEXT
# filters.TEXT | filters.CAPTION.TEXT
# filters.TEXT сработает, если .text есть и не None
# filters.CAPTION.TEXT сработает, если .caption есть, не None и не пустая строка
# Это то, что нужно.
# Но если пересылается сообщение без комментария, и в текущем сообщении нет .text и .caption,
# то ни один из этих фильтров не сработает.
# Поэтому, чтобы *поймать* пересланное сообщение, даже если в нём нет текста/подписи,
# нужно либо отдельно ловить filters.FORWARDED, либо ловить всё и проверять внутри.
# Однако, если в пересланном сообщении нет ссылки (ни в тексте, ни в подписи текущего сообщения),
# то обрабатывать его бессмысленно.
# Цель: обработать *любое* сообщение, в котором *может быть* ссылка на Telegram-пост.
# Это: сообщения с .text, сообщения с .caption, и, косвенно, пересланные, если они добавили .text или .caption.
# Поэтому фильтр TEXT | CAPTION.TEXT должен покрыть основные случаи.
# filters.TEXT | filters.CAPTION.TEXT
# filters.CAPTION.TEXT эквивалентно filters.CAPTION & filters.TEXT
# filters.CAPTION проверяет, что caption != None
# filters.TEXT проверяет, что caption != ""
# filters.TEXT проверяет .text
# filters.TEXT проверяет .text (для основного текста)
# filters.TEXT проверяет .caption (если CAPTION установлен)
# Нет, filters.TEXT проверяет только .text.
# filters.CAPTION проверяет наличие .caption (не None)
# filters.TEXT проверяет .text
# filters.CAPTION.TEXT = filters.CAPTION & lambda u: bool(u.effective_message.caption and u.effective_message.caption.strip())
# filters.TEXT = lambda u: bool(u.effective_message.text)
# filters.TEXT срабатывает, если .text есть и не пустой/не None
# filters.CAPTION.TEXT срабатывает, если .caption есть и не пустой/не None
# Это почти то, что нужно.
# filters.TEXT | filters.CAPTION.TEXT
# Это означает: сработает, если в сообщении есть .text или .caption (и caption не пустой)
# Это покроет:
# - Сообщения с текстом (filters.TEXT)
# - Сообщения с медиа и подписью (filters.CAPTION.TEXT)
# - Пересланные сообщения, если они добавили текст или подпись (тоже будет .text или .caption)
# Это должно сработать для большинства сценариев, включая пересылку с комментарием.
# Если пользователь *тупо* перешлёт пост без комментария, и в оригинале не было встроенной ссылки,
# то в текущем сообщении бота .text и .caption будут None или пустыми, и фильтр не сработает.
# Это ограничение Bot API.
# Итак, используем фильтр TEXT | CAPTION.TEXT

# --- ОБНОВЛЁННЫЕ ОБРАБОТЧИКИ ---

async def handle_text_or_caption(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обрабатывает сообщения с текстом или подписью, включая пересланные с комментарием."""
    user_id = update.effective_user.id
    init_user_excel(user_id)

    # Проверяем текст или подпись текущего сообщения
    text = update.message.text or update.message.caption or ""
    if text:
        link = extract_telegram_link(text)
        if link:
            # Сохраняем ссылку в контексте пользователя
            context.user_data['current_link'] = link
            await _present_link_options(update, context, link)
            return # Обработали, выходим

    # Если ссылка не найдена
    await update.message.reply_text("❌ Я не нашёл действительную ссылку на пост в Telegram в твоём сообщении. Отправь ссылку или перешли пост с комментарием содержащим ссылку.")


# Экспорт базы данных (остаётся без изменений)
async def export_database(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    excel_file = get_user_excel_file(user_id)
    if os.path.exists(excel_file):
        await update.message.reply_document(
            document=open(excel_file, 'rb'),
            filename=f'my_posts_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
    else:
        await update.message.reply_text("❌ Твоя база данных пуста. Добавь хотя бы один пост.")

# Статистика (остаётся без изменений)
async def stats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    excel_file = get_user_excel_file(user_id)
    if not os.path.exists(excel_file):
        await update.message.reply_text("📊 Твоя база данных пуста.")
        return
    
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    total = ws.max_row - 1
    
    statuses = {}
    for row in range(2, ws.max_row + 1):
        status = ws[f'C{row}'].value
        if status:
            statuses[status] = statuses.get(status, 0) + 1
    
    message = f"📊 Статистика *твоих* постов:\n\n"
    message += f"Всего постов: {total}\n\n"
    
    if statuses:
        message += "По статусам:\n"
        for status, count in statuses.items():
            message += f"• {status}: {count}\n"
    
    await update.message.reply_text(message)

# Основная функция
def main():
    global app

    TOKEN = os.getenv("BOT_TOKEN")
    
    if not TOKEN:
        logger.error("Требуется переменная окружения BOT_TOKEN")
        return

    app = Application.builder().token(TOKEN).job_queue(None).build()
    
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("export", export_database))
    app.add_handler(CommandHandler("stats", stats))
    app.add_handler(CallbackQueryHandler(button_handler))
    # Обновлённый обработчик: ловит сообщения с текстом или подписью
    app.add_handler(MessageHandler((filters.TEXT | filters.CAPTION.TEXT) & ~filters.COMMAND, handle_text_or_caption))
    
    logger.info("Бот запущен с поддержкой пересланных сообщений (если они содержат ссылку)!")
    app.run_polling(allowed_updates=Update.ALL_TYPES)

if __name__ == '__main__':
    main()